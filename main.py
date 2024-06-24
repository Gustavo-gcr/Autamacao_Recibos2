import os
import sys
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from docx import Document
from docx2pdf import convert
from datetime import datetime

# Aumentar a profundidade máxima de recursão
sys.setrecursionlimit(1500)

# Função para formatar a data para "dia/mês/ano"
def formatar_referencia_numerica(data):
    return data.strftime('%d/%m/%Y')

def formatar_referencia_numerica2(data):
    return data.strftime('%d-%m-%Y')

# Função para formatar a referência para "mês ano" em português brasileiro
def formatar_referencia(data):
    return data.strftime('%b %Y')

def selecionar_arquivo_excel():
    root = Tk()
    root.withdraw()  # Ocultar a janela principal
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    return file_path

def selecionar_pasta_salvar():
    root = Tk()
    root.withdraw()  # Ocultar a janela principal
    folder_path = filedialog.askdirectory()
    return folder_path

def selecionar_arquivo_word():
    root = Tk()
    root.withdraw()  # Ocultar a janela principal
    file_path = filedialog.askopenfilename(filetypes=[("Word Files", "*.docx")])
    return file_path

# Selecionar o arquivo Excel
arquivo_excel = selecionar_arquivo_excel()
if not arquivo_excel:
    print("Nenhum arquivo Excel selecionado. Encerrando o programa.")
    exit()

# Selecionar o arquivo Word
arquivo_word = selecionar_arquivo_word()
if not arquivo_word:
    print("Nenhum arquivo Word selecionado. Encerrando o programa.")
    exit()

# Selecionar a pasta de destino para salvar os documentos do Word
pasta_destino = selecionar_pasta_salvar()
if not pasta_destino:
    print("Nenhuma pasta de destino selecionada. Encerrando o programa.")
    exit()

# Carregar o arquivo do Excel
workbook = load_workbook(arquivo_excel)
sheet = workbook.active

# Iterar sobre as linhas do Excel, começando da segunda linha para ignorar o cabeçalho
for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    # Criar um novo documento do Word para cada linha
    document = Document(arquivo_word)
    
    # Extrair os valores da linha
    NOMEEMPREGADO, SETOR, CC, MOTIVO, DESTINO, ROTEIRO, TIPO, DATA1, DESCRICAO1, VALOR1, DOCUMENTO1, DATA2, DESCRICAO2, VALOR2, DOCUMENTO2, DATA3, DESCRICAO3, VALOR3, DOCUMENTO3,DATA4, DESCRICAO4, VALOR4, DOCUMENTO4, DATA5, DESCRICAO5, VALOR5, DOCUMENTO5  = row
    
    # Formatar as datas
   
    
    # Substituir os marcadores no documento do Word com os dados do Excel
    for paragraph in document.paragraphs:
        # Fazer as substituições de acordo com os marcadores no documento do Word
        paragraph.text = paragraph.text.replace('nome_empregado', f'{NOMEEMPREGADO}')
        paragraph.text = paragraph.text.replace('var_setor', SETOR)
        paragraph.text = paragraph.text.replace('centro_custo', CC)
        paragraph.text = paragraph.text.replace('var_motivo', MOTIVO)
        paragraph.text = paragraph.text.replace('var_destino', DESTINO)
        paragraph.text = paragraph.text.replace('var_roteiro', ROTEIRO)
        paragraph.text = paragraph.text.replace('tipo_viagem', str(TIPO))
        
        # TEMOS QUE ADD OS VALORES SEGUINTES OU CRIAR UM FOR PARA INTERAGIR COM AS VARIAVEIS
        if DATA1:
            data1_formatada = formatar_referencia_numerica(DATA1)
            paragraph.text = paragraph.text.replace('var_data1', str(data1_formatada))
            paragraph.text = paragraph.text.replace('descricao_gasto1', DESCRICAO1)
            VALOR1 = float(VALOR1)
            paragraph.text = paragraph.text.replace('var_valor1', f'R$ {VALOR1:.2f}')
            paragraph.text = paragraph.text.replace('var_documento1', DOCUMENTO1)
            
        if DATA2 != '':
            data2_formatada = formatar_referencia_numerica(DATA2)
            paragraph.text = paragraph.text.replace('var_data2', str(data2_formatada))
            paragraph.text = paragraph.text.replace('descricao_gasto2', DESCRICAO2)
            VALOR2 = float(VALOR2)
            paragraph.text = paragraph.text.replace('var_valor2', f'R$ {VALOR2:.2f}')
            paragraph.text = paragraph.text.replace('var_documento2', DOCUMENTO2)
            
            
        if DATA3 != None:
            data3_formatada = formatar_referencia_numerica(DATA3)
            paragraph.text = paragraph.text.replace('var_data3', str(data3_formatada))
            paragraph.text = paragraph.text.replace('descricao_gasto3', DESCRICAO3)
            VALOR3 = float(VALOR3)
            paragraph.text = paragraph.text.replace('var_valor3', f'R$ {VALOR3:.2f}')
            paragraph.text = paragraph.text.replace('var_documento3', DOCUMENTO3)
        else:
            # data3_formatada = formatar_referencia_numerica(DATA3)
            paragraph.text = paragraph.text.replace('var_data3','')
            paragraph.text = paragraph.text.replace('descricao_gasto3','' )
            # VALOR3 = float(VALOR3)
            paragraph.text = paragraph.text.replace('var_valor3','')
            paragraph.text = paragraph.text.replace('var_documento3', '')
            
        if DATA4 != None:
            data4_formatada = formatar_referencia_numerica(DATA4)
            paragraph.text = paragraph.text.replace('var_data4', str(data4_formatada))
            paragraph.text = paragraph.text.replace('descricao_gasto4', DESCRICAO4)
            VALOR4 = float(VALOR4)
            paragraph.text = paragraph.text.replace('var_valor4', f'R$ {VALOR4:.2f}')
            paragraph.text = paragraph.text.replace('var_documento4', DOCUMENTO4)
        else:
            paragraph.text = paragraph.text.replace('var_data4', '')
            paragraph.text = paragraph.text.replace('descricao_gasto4', '')
            # VALOR4 = float(VALOR4)
            paragraph.text = paragraph.text.replace('var_valor4','')
            paragraph.text = paragraph.text.replace('var_documento4', '')
            
        if DATA5 != None:
            data5_formatada = formatar_referencia_numerica(DATA5)
            paragraph.text = paragraph.text.replace('var_data5', str(data5_formatada))
            paragraph.text = paragraph.text.replace('descricao_gasto5', DESCRICAO5)
            VALOR5 = float(VALOR5)
            paragraph.text = paragraph.text.replace('var_valor5', f'R$ {VALOR5:.2f}')
            paragraph.text = paragraph.text.replace('var_documento5', DOCUMENTO5)
        else:
            paragraph.text = paragraph.text.replace('var_data5','')
            paragraph.text = paragraph.text.replace('descricao_gasto5', '')
            # VALOR5 = float(VALOR5)
            paragraph.text = paragraph.text.replace('var_valor5', '')
            paragraph.text = paragraph.text.replace('var_documento5','')
            
    # Salvar o documento do Word modificado na pasta de destino
    nome_arquivo = f'Recibo - {NOMEEMPREGADO}_{CC}_{DESTINO}.docx'
    caminho_docx = os.path.join(pasta_destino, nome_arquivo)
    document.save(caminho_docx)
    
    # Gerar o arquivo PDF correspondente
    convert(caminho_docx)

print("Documentos criados com sucesso.")


